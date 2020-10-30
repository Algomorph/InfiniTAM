import argparse
import pandas as pd
import numpy as np
import pickle
import sys
from tuner import truncation_distance_factor_list, voxel_size_list, switch_set_names, \
    warp_length_termination_threshold_type_names
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Side, Border

PROGRAM_SUCCESS = 0
PROGRAM_FAILURE = -1


def merge_dictionaries(dict1, dict2):
    res = {**dict1, **dict2}
    return res


def first_item_or_zero_if_empty(some_list):
    return 0 if len(some_list) == 0 else some_list[0]


def dict_of_lists_to_dict_of_first_item_or_zero(dict_in):
    dict_out = {}
    for key, value in dict_in.items():
        dict_out[key] = first_item_or_zero_if_empty(value)
    return dict_out


def get_keys_with_dots(string_keyed_dict):
    return [key for key in string_keyed_dict.keys() if '.' in key]


def get_associated_key_without_dot(key):
    return key if '.' not in key else key.split('.')[1]


def remove_key_hierarchy(dict_with_key_hierarchy, required_keys):
    dict_out = {}
    for key, value in dict_with_key_hierarchy.items():
        new_value = first_item_or_zero_if_empty(value)
        new_key = get_associated_key_without_dot(key)
        if new_key not in dict_out or dict_out[new_key] == 0:
            dict_out[new_key] = new_value
    for key in required_keys:
        if key not in dict_out or dict_out[key] is None:
            dict_out[key] = 0
    return dict_out


def main() -> int:
    parser = argparse.ArgumentParser(
        description='Convert a pickled hyperopt trials object from the tuner to Excel spreadsheet')
    parser.add_argument('-to', '--trials_object', default="tuning_table.hyperopt", type=str,
                        help='Path to the pickled trials object resulting from the tuner.')
    parser.add_argument('-o', '--output', default="tuning_table.xlsx", type=str,
                        help='Path to the pickled trials object resulting from the tuner.')

    args = parser.parse_args()
    trials = pickle.load(open(args.trials_object, "rb"))
    required_keys = [get_associated_key_without_dot(key) for key in
                     get_keys_with_dots(trials.trials[0]["misc"]["vals"])]

    to_save = [merge_dictionaries(remove_key_hierarchy(trial["misc"]["vals"], required_keys), trial["result"])
               for trial in trials.trials]

    df = pd.DataFrame(to_save)
    df["voxel_size"] = np.array(voxel_size_list)[df["voxel_size"].to_numpy()]
    df["truncation_distance"] = np.array(truncation_distance_factor_list)[df["truncation_distance_factor"].to_numpy()] \
                                * df["voxel_size"]
    df = df.drop(columns=["truncation_distance_factor"])
    df["switch_set"] = np.array(switch_set_names)[df["switch_set"].to_numpy()]
    df["warp_length_termination_threshold_type"] = np.array(warp_length_termination_threshold_type_names)[
        df["warp_length_termination_threshold_type"].to_numpy()]

    writer = pd.ExcelWriter(args.output)

    df = df.sort_values(by='loss')
    df.to_excel(writer, 'Trials')

    writer.save()

    workbook = openpyxl.open(args.output)
    trials_worksheet = workbook.worksheets[0]
    column_widths = []
    for row in trials_worksheet.iter_rows():
        for i_column, cell in enumerate(row):
            try:
                column_widths[i_column] = max(column_widths[i_column], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    font = Font(name="Liberation Sans", bold=False, size=12)
    header_fill = PatternFill(start_color="00B2B2B2", end_color="00B2B2B2", fill_type='solid')
    odd_column_fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type='solid')
    even_column_fill = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type='solid')
    solid_border = Side(style="thin", color="00323232")
    border_style = Border(bottom=solid_border)

    for i_column, column_width in enumerate(column_widths):
        trials_worksheet.column_dimensions[get_column_letter(i_column + 1)].width = column_width

    i_row = 0
    for row in trials_worksheet.iter_rows(1, len(df) + 1, 1, len(df.columns) + 1):
        if i_row == 0:
            current_odd_fill = header_fill
            current_even_fill = header_fill
        else:
            current_odd_fill = odd_column_fill
            current_even_fill = even_column_fill

        for i_column in range(0, len(row)):
            row[i_column].font = font
            row[i_column].border = border_style
            row[i_column].fill = current_odd_fill if i_column % 2 == 0 else current_even_fill
        i_row += 1

    workbook.save(args.output)
    workbook.close()

    return PROGRAM_SUCCESS


if __name__ == "__main__":
    sys.exit(main())
